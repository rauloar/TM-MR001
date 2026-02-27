from PySide6.QtWidgets import QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QFileDialog, QPushButton
from attendance.models import AttendanceLog, User
import pandas as pd

class MovementsTab(QWidget):
    def __init__(self, session):
        super().__init__()
        self.session = session

        from PySide6.QtWidgets import QHBoxLayout, QLabel, QComboBox, QDateEdit, QCheckBox
        from PySide6.QtCore import QDate
        layout = QVBoxLayout()

        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Fecha inicio:"))
        self.date_start = QDateEdit()
        self.date_start.setCalendarPopup(True)
        self.date_start.setDate(QDate.currentDate())
        filter_layout.addWidget(self.date_start)

        filter_layout.addWidget(QLabel("Fecha fin:"))
        self.date_end = QDateEdit()
        self.date_end.setCalendarPopup(True)
        self.date_end.setDate(QDate.currentDate())
        filter_layout.addWidget(self.date_end)

        filter_layout.addWidget(QLabel("Usuario(s):"))
        self.user_selector = QComboBox()
        self.user_selector.setEditable(True)
        self.user_selector.addItem("Todos")
        users = self.session.query(User).filter(User.is_active==True).all()
        for user in users:
            display = (user.first_name or "") + " " + (user.last_name or "")
            self.user_selector.addItem(f"{user.identifier} - {display.strip()}", user.identifier)
        filter_layout.addWidget(self.user_selector)


        layout.addLayout(filter_layout)

        from PySide6.QtWidgets import QSizePolicy
        from PySide6.QtWidgets import QHeaderView
        from PySide6.QtCore import Qt
        self.table = QTableWidget(self)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        layout.addWidget(self.table, stretch=1)

        from PySide6.QtWidgets import QStyle
        style = self.style()
        self.btn_refresh = QPushButton("Actualizar")
        self.btn_refresh.setIcon(style.standardIcon(QStyle.StandardPixmap.SP_BrowserReload))
        self.btn_refresh.clicked.connect(self.load_data)
        layout.addWidget(self.btn_refresh)

        self.btn_export = QPushButton("Exportar CSV")
        self.btn_export.setIcon(style.standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        self.btn_export.clicked.connect(self.export_csv)
        layout.addWidget(self.btn_export)

        self.btn_export_excel = QPushButton("Exportar Excel")
        self.btn_export_excel.setIcon(style.standardIcon(QStyle.StandardPixmap.SP_DriveHDIcon))
        self.btn_export_excel.clicked.connect(self.export_excel)
        layout.addWidget(self.btn_export_excel)

        self.setLayout(layout)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.load_data()

    def update_user_selector(self):
        self.user_selector.clear()
        self.user_selector.addItem("Todos")
        users = self.session.query(User).filter(User.is_active==True).all()
        for user in users:
            display = (user.first_name or "") + " " + (user.last_name or "")
            self.user_selector.addItem(f"{user.identifier} - {display.strip()}", user.identifier)

    def load_data(self):
        from sqlalchemy.orm import joinedload
        from datetime import date
        start = self.date_start.date().toPython()
        end = self.date_end.date().toPython()
        user_value = self.user_selector.currentData()
        query = self.session.query(AttendanceLog, User).join(User, AttendanceLog.raw_identifier == User.identifier)
        query = query.filter(User.is_active == True)
        # Solo filtrar por usuario si no es 'Todos'
        if user_value and user_value != None and self.user_selector.currentText() != "Todos":
            query = query.filter(User.identifier == user_value)
        if start:
            query = query.filter(AttendanceLog.date >= start)
        if end:
            query = query.filter(AttendanceLog.date <= end)
        logs = query.all()
        self.table.setRowCount(len(logs))
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Identificador", "Nombre", "Fecha", "Hora", "Movimiento"])
        for row, (log, user) in enumerate(logs):
            movimiento = "ENTRADA" if log.mark_type == 0 else "SALIDA"
            if user.first_name or user.last_name:
                nombre = (user.first_name or "") + " " + (user.last_name or "")
            else:
                nombre = log.raw_identifier
            fecha_str = log.date.strftime("%d/%m/%Y") if log.date else ""
            hora_str = log.time.strftime("%H:%M") if log.time else ""
            self.table.setItem(row, 0, QTableWidgetItem(log.raw_identifier))
            self.table.setItem(row, 1, QTableWidgetItem(nombre))
            self.table.setItem(row, 2, QTableWidgetItem(fecha_str))
            self.table.setItem(row, 3, QTableWidgetItem(hora_str))
            self.table.setItem(row, 4, QTableWidgetItem(movimiento))

    def export_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "Guardar CSV", "", "CSV Files (*.csv)")
        if path:
            data = []
            selected_rows = self.table.selectionModel().selectedRows()
            if selected_rows:
                for index in selected_rows:
                    row = index.row()
                    log_id = self.table.item(row, 0).text()
                    nombre = self.table.item(row, 1).text()
                    fecha_str = self.table.item(row, 2).text()
                    hora_str = self.table.item(row, 3).text()
                    movimiento = self.table.item(row, 4).text()
                    data.append([log_id, nombre, fecha_str, hora_str, movimiento])
            else:
                for row in range(self.table.rowCount()):
                    log_id = self.table.item(row, 0).text()
                    nombre = self.table.item(row, 1).text()
                    fecha_str = self.table.item(row, 2).text()
                    hora_str = self.table.item(row, 3).text()
                    movimiento = self.table.item(row, 4).text()
                    data.append([log_id, nombre, fecha_str, hora_str, movimiento])
            df = pd.DataFrame(data, columns=["Identificador", "Nombre", "Fecha", "Hora", "Movimiento"])
            df.to_csv(path, index=False)

    def export_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", "", "Excel Files (*.xlsx)")
        if path:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            data = []
            selected_rows = self.table.selectionModel().selectedRows()
            if selected_rows:
                for index in selected_rows:
                    row = index.row()
                    log_id = self.table.item(row, 0).text()
                    nombre = self.table.item(row, 1).text()
                    fecha_str = self.table.item(row, 2).text()
                    hora_str = self.table.item(row, 3).text()
                    movimiento = self.table.item(row, 4).text()
                    data.append([log_id, nombre, fecha_str, hora_str, movimiento])
            else:
                for row in range(self.table.rowCount()):
                    log_id = self.table.item(row, 0).text()
                    nombre = self.table.item(row, 1).text()
                    fecha_str = self.table.item(row, 2).text()
                    hora_str = self.table.item(row, 3).text()
                    movimiento = self.table.item(row, 4).text()
                    data.append([log_id, nombre, fecha_str, hora_str, movimiento])
            # Crear workbook y hoja
            wb = openpyxl.Workbook()
            ws = wb.active
            import os
            sheet_name = os.path.splitext(os.path.basename(path))[0]
            ws.title = sheet_name
            headers = ["Identificador", "Nombre", "Fecha", "Hora", "Movimiento"]
            # Encabezados en negrita
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Datos
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max(15, max_length + 2)
            wb.save(path)